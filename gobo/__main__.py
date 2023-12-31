from asyncio import run
from collections.abc import Callable, Coroutine
from contextlib import suppress
from functools import wraps
from typing import Any, ParamSpec, TypeVar

import click
from openpyxl import Workbook

from .database import db
from .types import SpotID

P = ParamSpec("P")
T = TypeVar("T")


def run_decorator(f: Callable[P, Coroutine[Any, Any, T]]) -> Callable[P, T]:
    @wraps(f)
    def wrapped(*args: P.args, **kwargs: P.kwargs) -> T:
        return run(f(*args, **kwargs))

    return wrapped


@click.group
def main() -> None:
    ...


CLEARED = "A"
NAME = "B"
MUNICIPALITY = "C"


@main.command
@click.argument("output", type=click.Path(dir_okay=False))
@run_decorator
async def excel(
    output: str,
) -> None:
    wb = Workbook()

    spot_sheet = wb.create_sheet("スポット")
    spot_sheet[f"{CLEARED}1"] = "達成"
    spot_sheet[f"{NAME}1"] = "名前"
    spot_sheet[f"{MUNICIPALITY}1"] = "市町村"

    for i, spot_id in enumerate(db.spots, start=2):
        spot_sheet[f"{CLEARED}{i}"] = False
        spot_sheet[f"{NAME}{i}"].value = db.spot_name(spot_id).replace("\u3000", " ")
        spot_sheet[f"{MUNICIPALITY}{i}"] = scraping_address(spot_id)
        spot_sheet[f"D{i}"].value = "リンク(GoGo房総)"
        spot_sheet[f"D{i}"].hyperlink = f"https://platinumaps.jp/d/gogo-boso?s={spot_id}"
        with suppress(ValueError):
            spot_sheet[f"E{i}"].hyperlink = db.spot_uri(spot_id)
            spot_sheet[f"E{i}"].value = "リンク(施設)"

    spot_sheet.auto_filter.ref = spot_sheet.dimensions

    spot_clear_range = f"スポット!${CLEARED}${1+1}:${CLEARED}${1+len(db.spots)}"
    spot_area_range = f"スポット!${MUNICIPALITY}${1+1}:${MUNICIPALITY}${1+len(db.spots)}"

    total_sheet = wb.create_sheet("集計")
    total_sheet["A1"] = "市町村"
    total_sheet["B1"] = "達成数"
    total_sheet["C1"] = "総数"
    total_sheet["D1"] = "達成率"
    for i, municipality_id in enumerate(db.municipalities, start=2):
        total_sheet[f"A{i}"] = db.municipality_name(municipality_id)
        total_sheet[f"B{i}"] = (
            f"=COUNTIFS({spot_area_range}, "
            f'"*{db.municipality_name(municipality_id)}*", {spot_clear_range}, TRUE)'
        )
        total_sheet[
            f"C{i}"
        ] = f'=COUNTIFS({spot_area_range}, "*{db.municipality_name(municipality_id)}*")'
        total_sheet[f"D{i}"] = f"=100 * $B${i} / $C${i}"

    wb.remove(wb.worksheets[0])
    wb.save(output)


def scraping_address(spot_id: SpotID) -> str:
    c = db.connection.cursor()
    c.execute(
        """
SELECT spot_address
FROM spot_addresses
WHERE spot_id = ?
        """,
        (spot_id,),
    )
    (address,) = c.fetchone()

    if address == "印旛郡酒々井町本佐倉・佐倉市大佐倉":
        return "酒々井町;佐倉市"
    if address == "夷隅郡大多喜町粟又～市原市朝生原":
        return "大多喜町;市原市"
    if address == "市原市・長生郡長柄町":
        return "市原市;長柄町"

    names = [
        name
        for name in map(db.municipality_name, db.municipalities)
        if name in address.replace("ケ", "ヶ").replace("舘山", "館山")
    ]

    if len(names) == 1:
        return names[0]
    else:
        raise ValueError(address, names)


if __name__ == "__main__":
    main()
